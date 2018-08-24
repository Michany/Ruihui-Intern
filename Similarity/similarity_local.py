# -*- coding: utf-8 -*-
"""
Created on Sat Jul 21 22:59:47 2018

@author: 70242

## 运行说明  

- 含excel转换为pdf的功能  
  需要win32com包，请在命令行中输入 pip install pywin32 安装。

- 请保证.py程序运行目录下包含有 sh.xlsx 历史数据文件。

- 请保证.py程序运行目录下包含有 up.png, down.png 两个图片文件。

"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import tushare as ts 
from win32com import client

TARGET_DAYS = [5, 20]
FREQUENCY = 15
TARGET_WINDOWS = [int(i*4*60/FREQUENCY) for i in TARGET_DAYS]
INDEX_SYMBOL = '000016.SH'

#%% 获取历史行情数据
t = pd.read_excel('sh.xlsx', dtype={'ddate':'datetime64'})
t.rename(columns={'ddate':'date', 'sclose':'close'}, inplace=True)
t.set_index('date', inplace=True)
t = t.close

now = ts.get_hist_data(code='sh', start=None, end=None,
                       ktype='5', retry_count=3, pause=0.001)
now.sort_values(by='date', ascending=True, inplace=True)

print('Data prepared!')

#%% 计算相关系数
len_t, len_now = len(t), len(now)
max_corr, max_index = 0, -1
corr = list()
for i in range(0, len_t-len_now):
    temp_corr = np.corrcoef(x=t[i:i+len_now], y=now['close'])
    corr.append(temp_corr[1])
corr = pd.DataFrame(corr)
ans = corr[corr[0] > 0.75]
print('Correlation done')

#%% 对满足条件的历史序列分组，如果index连续则视为一组
group_list = []
group, last_i = 0, 0
for i in ans.index:
    if i-last_i != 1:
        group += 1
        last_i = i
    else:
        last_i += 1
    group_list.append(group)
ans = pd.DataFrame({0: list(ans[0]), 1: group_list}, index=ans.index)
ans_groups = ans.groupby(1)

#%% 计算概率
def cal_prob(d: int):
    global forward_length, prob_up, prob_down
    forward_length = TARGET_WINDOWS[d]
    index_max_corr = [
        int(ans_groups.get_group(j + 1).idxmax()[0]) for j in range(ans_groups.size().size)
    ]
    prob_up = sum(t[i] < t[i + len_now + forward_length]
                  for i in index_max_corr) / ans_groups.size().size
    prob_down = 1 - prob_up
    print('{} days:\nprob_up: {};\n prob_down: {};'.format(
        TARGET_DAYS[d], prob_up, prob_down))
    return prob_up, prob_down

cal_prob(1)
index_max_corr = [int(ans_groups.get_group(j+1).idxmax()[0]) for j in range(ans_groups.size().size)]

if prob_up > prob_down:
    for i in index_max_corr:
        if t[i] < t[i+len_now+forward_length]:
            ans.loc[i, 'more_likely'] = 1
else:
    for i in index_max_corr:
        if t[i] > t[i+len_now+forward_length]:
            ans.loc[i, 'more_likely'] = 1

best_history_index = ans.where(ans.more_likely != 0).idxmax()[0]


#%% 调整坐标及缩放
def adjust_axis_limit():
    global future, curt, a, b
    i = best_history_index

    future = dict()
    future['upper'] = t[i:i+len_now+forward_length].max()
    future['lower'] = t[i:i+len_now+forward_length].min()
    future['mid'] = (future['upper']+future['lower'])/2
    future['width'] = future['upper']-future['lower']
    future['average'] = t[i:i+len_now+forward_length].mean()
    future['upward'] = future['upper']-future['average']
    future['downward'] = future['average']-future['lower']

    curt = dict()
    curt['upper'] = now['close'].max()
    curt['lower'] = now['close'].min()
    curt['average'] = now['close'].mean()
    curt['width'] = curt['upper']-curt['lower']
    curt['mid'] = (curt['upper']+curt['lower'])/2

    a = (curt['lower']-curt['width']*future['downward'] /
         future['width']*3)*future['average']/curt['average']
    b = (curt['upper']+curt['width']*future['upward'] /
         future['width']*3)*future['average']/curt['average']
adjust_axis_limit()

#%% 画图
fig = plt.figure()
fig.set_size_inches(16, 12)

# 放置主坐标轴（左轴），用于标示当前行情序列
ax = fig.add_subplot(111)
ax.plot(range(t.index[best_history_index].toordinal(), t.index[best_history_index].toordinal() + len_now), 
        now['close'], linewidth=3, label='current', c='#007ac6')
plt.ylim((curt['mid'] - future['downward'] * 1, curt['mid'] + future['upward'] * 1)) #调整刻度，更美观地对其
plt.grid(color='#007ac6', alpha=0.7, axis='y',linestyle = 'dashed')
plt.yticks(fontsize=12)
ax.spines['top'].set_visible(False)
ax.spines['bottom'].set_visible(False)
ax.spines['left'].set_visible(False)
ax.spines['right'].set_visible(False)

# 放置次坐标轴（右轴），用于标示历史行情序列
ax2 = ax.twinx()
history = t[best_history_index:best_history_index + len_now + forward_length]
ax2.plot(
    range(t.index[best_history_index].toordinal(), t.index[best_history_index].toordinal()+len_now+forward_length), 
    history, linewidth=3, label='history', c='#f08600')
ax2.spines['top'].set_visible(False)
ax2.spines['bottom'].set_visible(False)
ax2.spines['left'].set_visible(False)
ax2.spines['right'].set_visible(False)
plt.ylim((a, b))
plt.axis('off')

# 标记x轴
plt.xticks()

fig.savefig(r'curve.png', dpi=144, bbox_inches='tight')
fig.savefig(r'curve [矢量图].pdf', dpi=144, bbox_inches='tight')


#%% 写入excel并设置样式
def output_to_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, numbers, Side, Alignment
    from openpyxl.drawing.image import Image

    #%% 定义样式
    font_col = Font(name='dengxian', size=20, bold=True, color='FFFFFFFF')
    font_panzheng = Font(name='dengxian', size=19, bold=True, color='FFF08600')
    font_prob = Font(name='dengxian', size=19, bold=False, color='FF007AC6')

    fill_col = PatternFill(
        fill_type='solid', start_color='FF0070C6', end_color='FF0070C6')
    fill_row = PatternFill(
        fill_type='solid', start_color='FFE4F3FC', end_color='FFE4F3FC')
    fill_prob = PatternFill(
        fill_type='solid', start_color='FFF7F7F7', end_color='FFF7F7F7')

    format_percent = numbers.BUILTIN_FORMATS[9]

    border_corner1 = Border(left=Side(border_style='thick', color='FF999999'),
                            right=Side(border_style='thin', color='FFFFFFFF'),
                            top=Side(border_style='thick', color='FF999999'),
                            bottom=Side(border_style='thin', color='FFFFFFFF'))
    border_corner2 = Border(left=Side(border_style='thin', color='FFFFFFFF'),
                            right=Side(border_style='thick', color='FF999999'),
                            top=Side(border_style='thick', color='FF999999'),
                            bottom=Side(border_style='thin', color='FFFFFFFF'))
    border_corner3 = Border(left=Side(border_style='thin', color='FFFFFFFF'),
                            right=Side(border_style='thick', color='FF999999'),
                            top=Side(border_style='thin', color='FFFFFFFF'),
                            bottom=Side(border_style='thick', color='FF999999'))
    border_corner4 = Border(left=Side(border_style='thick', color='FF999999'),
                            right=Side(border_style='thin', color='FFFFFFFF'),
                            top=Side(border_style='thin', color='FFFFFFFF'),
                            bottom=Side(border_style='thick', color='FF999999'))
    border_up = Border(left=Side(border_style='thin', color='FFFFFFFF'),
                       right=Side(border_style='thin', color='FFFFFFFF'),
                       top=Side(border_style='thick', color='FF999999'),
                       bottom=Side(border_style='thin', color='FFFFFFFF'))
    border_left = Border(left=Side(border_style='thick', color='FF999999'),
                         right=Side(border_style='thin', color='FFFFFFFF'),
                         top=Side(border_style='thin', color='FFFFFFFF'),
                         bottom=Side(border_style='thin', color='FFFFFFFF'))
    border_right = Border(left=Side(border_style='thin', color='FFFFFFFF'),
                          right=Side(border_style='thick', color='FF999999'),
                          top=Side(border_style='thin', color='FFFFFFFF'),
                          bottom=Side(border_style='thin', color='FFFFFFFF'))
    border_down = Border(left=Side(border_style='thin', color='FFFFFFFF'),
                         right=Side(border_style='thin', color='FFFFFFFF'),
                         top=Side(border_style='thin', color='FFFFFFFF'),
                         bottom=Side(border_style='thick', color='FF999999'))
    border_inside = Border(left=Side(border_style='thin', color='FFFFFFFF'),
                           right=Side(border_style='thin', color='FFFFFFFF'),
                           top=Side(border_style='thin', color='FFFFFFFF'),
                           bottom=Side(border_style='thin', color='FFFFFFFF'))
    alignment = Alignment(horizontal='center', vertical='center',
                          text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
          
    wb = Workbook()
    ws = wb.active

    prob_up_5, prob_down_5 = cal_prob(0)
    prob_up_20, prob_down_20 = cal_prob(1)

    ws.append(['5天','概率', '20天', '概率'])
    ws.append(['', prob_up_5, '', prob_up_20])
    ws.append(['盘整', 0, '盘整', 0])
    ws.append(['', prob_down_5, '', prob_down_20])

    up, up2 = Image('up.png'), Image('up.png')
    down, down2 = Image('down.png'), Image('down.png')
    ws.add_image(up, 'A2')
    ws.add_image(up2, 'C2')
    ws.add_image(down, 'A4')
    ws.add_image(down2, 'C4')    
    #curve=Image('.\curve.png')
    #ws.add_image(curve, 'A6')

    for c in ['A','B','C','D']:
        ws.column_dimensions[c].width = 20
    for r in range(1,5):
        ws.row_dimensions[r].height = 35

    for row in range(1,5):
        for col in range(4):
            ws[row][col].border = border_inside
    ws[1][0].border, ws[1][3].border, ws[4][3].border, ws[4][0].border = border_corner1, border_corner2, border_corner3, border_corner4
    ws[1][1].border, ws[1][2].border = border_up, border_up
    ws[4][1].border, ws[4][2].border = border_down, border_down
    ws[2][0].border, ws[3][0].border = border_left, border_left
    ws[2][3].border, ws[3][3].border = border_right, border_right

    for row in range(1,5):
        for col in [0,2]:
            ws[row][col].fill = fill_row
            ws[row][col].font = font_col
            ws[row][col].alignment = alignment
        for col in [1,3]:
            ws[row][col].fill = fill_prob
            ws[row][col].font = font_prob
            ws[row][col].number_format = format_percent
            ws[row][col].alignment = alignment
    for col in range(4):
        ws[1][col].fill = fill_col
        ws[1][col].font = font_col
    ws[3][0].font, ws[3][2].font = font_panzheng, font_panzheng

    wb.save(INDEX_SYMBOL+"_Probability.xlsx")

def excel_to_pdf():
    xlApp = client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(INDEX_SYMBOL+"_Probability.xlsx")
    books.ExportAsFixedFormat(0, INDEX_SYMBOL+"_Probability.pdf")
    xlApp.Quit()

excel_to_pdf()
