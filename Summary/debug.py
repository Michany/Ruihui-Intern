import pandas as pd
import openpyxl
def convert_types(table, nrows, ncols):
    def delete(original, pos):
        '''Inserts new inside original at pos.'''
        new = original[:pos] + original[pos+1:]
        return new

    def initial(sin):
        waiting = []
        count = 0
        for i in range(len(sin)):
            if sin[i] in ',¥':
                waiting.append(i)
        s = sin
        for i in waiting:
            s = delete(s, (i - count))
            count += 1
        return s

    abnormal = False
    for i in [4,7]:
        record = [0]
        for j in range(5, nrows+1):
            try:
                record.append(float(initial(table[j][i].value)))
            except:
                abnormal = True
                record.append(table[j][i].value)
        print('正在转换数据类型', t.keys()[i])
        t[t.keys()[i]] = record
    if abnormal:
        t.drop(t.index[[nrows-5, nrows-4]], inplace=True)  # 如果最后有未清仓 则删掉最后两行

try:
    # 读取excel文件，sheet_name='交易列表',header=2（标题位于第三行）

    t = pd.read_excel(r"E:\RH\excel\IC1Year.xlsx", sheet_name='交易列表', header=2, usecols='A:I' )
    w = openpyxl.load_workbook(r"E:\RH\excel\IC1Year.xlsx")
    table = w.worksheets[2]
    nrows = table.max_row
    ncols = table.max_column
except:
    print('找不到文件/读取文件错误')
print('正在初始化...')    
convert_types(table, nrows, ncols)

# %% 统计回撤
def drawdown_by_time(t: pd.DataFrame):
    '''
    计算内容：最大回撤比例，累计收益率  
    计算方式：单利  
    返回结果：最大回撤率，开始日期，结束日期，总收益率，年化收益，年化回撤  
    '''
    t['Capital'] = float(t['价格'][1])+t['Net Profit - Cum Net Profit']

    yearly_drawdown = dict()
    
    t['Year']=pd.Series(t['Date/Time'][i].year for i in range(len(t)))
    t_group=t.groupby('Year')
    year_groups=[t_group.get_group(i) for i in t_group.groups.keys()]
    for year_group in year_groups:
        max_draw_down, temp_max_value = 0, 0
        start_date, end_date, current_start_date = 0, 0, 0
        continous = False # 是否连续
        for i in year_group.index:
            if year_group['#'][i]>0: continue
            if temp_max_value < year_group['Capital'][i]:
                current_start_date = year_group['Date/Time'][i]
            temp_max_value = max(temp_max_value, year_group['Capital'][i])
            if max_draw_down>year_group['Capital'][i]/temp_max_value-1:
                if not continous: 
                    continous = True
                max_draw_down = year_group['Capital'][i]/temp_max_value-1
            else:
                if continous:
                    continous = False
                    start_date = current_start_date
                    end_date = year_group['Date/Time'][i]
        yearly_drawdown[year_group['Year'][i]] = max_draw_down, start_date, end_date            

    yearly_return = dict() # 记录年收益率
    max_draw_down, temp_max_value = 0, 0
    start_date, end_date, current_start_date = 0, 0, 0
    continous = False # 是否连续

    for i in range(2, len(t),2): # 偶数行的数据
        if temp_max_value < t['Capital'][i-2]:
            current_start_date = t['Date/Time'][i]
        temp_max_value = max(temp_max_value, t['Capital'][i-2])
        if max_draw_down>t['Capital'][i]/temp_max_value-1:
            if not continous: 
                continous = True
            max_draw_down = t['Capital'][i]/temp_max_value-1
        else:
            if continous:
                continous = False
                start_date = current_start_date
                end_date = t['Date/Time'][i]

        # 统计年收益率
        year = t['Date/Time'][i].year
        yearly_return[year] = t['Net Profit - Cum Net Profit'][i]
    
    total_return = t['Capital'][i]/t['Capital'][0]-1
    yearly_return = pd.Series(yearly_return) / t['Capital'][0]
    first_year = t['Date/Time'][1].year, yearly_return[t['Date/Time'][1].year]
    yearly_return = yearly_return.diff()
    yearly_return[first_year[0]] = first_year[1]
    return max_draw_down, start_date, end_date, total_return, pd.Series(yearly_return), yearly_drawdown

drawdown_by_time(t)