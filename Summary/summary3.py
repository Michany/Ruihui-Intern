# -*- coding: utf-8 -*-
import time

import openpyxl
import pandas as pd
import numpy as np


def convert_types(table, nrows, ncols):
    global t

    def delete(original, pos):
        '''Inserts new inside original at pos.'''
        new = original[:pos] + original[pos+1:]
        return new

    def initial(sin):
        waiting = []
        count = 0
        for i in range(len(sin)):
            if sin[i] in ',¥':
                waiting.append(i)
        s = sin
        for i in waiting:
            s = delete(s, (i - count))
            count += 1
        return s

    for i in [4, 6, 7, 9]:
        record = [0]
        for j in range(5, nrows+1):
            try:
                record.append(float(initial(table[j][i].value)))
            except:
                record.append(table[j][i].value)
        print('正在转换数据类型', t.keys()[i])
        t[t.keys()[i]] = record


class Sum_class():
    '''
    一个包含了signal的统计信息的类
    --------------------------

    包含的属性如下

    name, volume, profit

    需要逐个查看signal对象时，可以使用

    >>> s.signal_class_list
    '''

    def __init__(self, input_name):
        self.name = input_name
        self.volume = dict()
        self.sum_volume = 0
        self.profit = dict()
        self.sum_profit = 0
        self.timeseries = dict()
        self.win = dict()   #胜率
        self.win_count = dict()
        self.loss = dict()  #赔率
        self.loss_count = dict()

    def check(self):
        '''
        统计该Signal下的各类信息
        '''
        total_row = len(t)
        for i in range(total_row):
            is_active = t.loc[i]["类型"] in ["买入", "卖空"]  # 属于主动买入或者卖空操作
            if t.loc[i]['Signal'] == self.name and is_active:
                # 记录volume
                j = i+1  # j为行控制变量，一直向下查找
                while t.loc[j]['#'] != t.loc[i]['#']+1:  # 当记录还没有跳到下一条时
                    # 记录总收益
                    self.sum_volume += t.loc[i]['Shares/Ctrts/Units - Profit/Loss']
                    self.volume[t.loc[j]['Signal']] = self.volume.get(
                        t.loc[j]['Signal'], 0) + t.loc[i]['Shares/Ctrts/Units - Profit/Loss']

                    # 记录总收益
                    self.sum_profit += t.loc[i]['Net Profit - Cum Net Profit']
                    self.profit[t.loc[j]['Signal']] = self.profit.get(
                        t.loc[j]['Signal'], 0) + t.loc[j]['Shares/Ctrts/Units - Profit/Loss']
                    
                    #记录胜率 赔率
                    if t.loc[i]['Net Profit - Cum Net Profit'] >= 0:
                        self.win[t.loc[j]['Signal']] = self.win.get(t.loc[j]['Signal'], 0) + t.loc[i]['Net Profit - Cum Net Profit']
                        self.win_count[t.loc[j]['Signal']] = self.win_count.get(t.loc[j]['Signal'], 0) + 1
                    else:
                        self.loss[t.loc[j]['Signal']] = self.loss.get(t.loc[j]['Signal'], 0) + t.loc[i]['Net Profit - Cum Net Profit']
                        self.loss_count[t.loc[j]['Signal']] = self.loss_count.get(t.loc[j]['Signal'], 0) + 1

                    # 记录收益对应的时间
                    self.timeseries[t.loc[j]['Date/Time']] = self.timeseries.get(
                        t.loc[j]['Date/Time'], 0) + t.loc[i]['% Profit']
                    j += 1
                    if j >= total_row:
                        break

        #胜率
        self.win_ratio = len(self.win)/(len(self.win)+len(self.loss))
def read_files(filepath):
    '''
    t 为pandas载入的DataFrame 

    w 为openpyxl载入的原excel文件，用于转换数据类型

    f 为文件路径
    '''
    global t, w, f
    try:
        # 读取excel文件，sheet_name='交易列表',header=2（标题位于第三行）
        f = filepath
        t = pd.read_excel(filepath, sheet_name='交易列表', header=2)
        w = openpyxl.load_workbook(filepath)
        table = w.worksheets[2]
        nrows = table.max_row
        ncols = table.max_column
    except:
        print('找不到文件')
    convert_types(table, nrows, ncols)


def init_signal_class():
    '''
    获取所有主动操作的signal names。

    依次初始化各个signal:Sum_class对象。
    '''
    global signal_class_list

    # 统计所有属于主动买入的Signal类型
    buy_signal_types = t[t[u"类型"] == u"买入"].groupby('Signal')
    # 统计所有属于主动卖空的Signal类型
    short_signal_types = t[t['类型'] == '卖空'].groupby('Signal')
    buy_signals_name_list = buy_signal_types.size()._index
    short_signals_name_list = short_signal_types.size()._index

    signal_class_list = []
    for signal in buy_signals_name_list:
        signal_class_list.append(Sum_class(signal))
    for signal in short_signals_name_list:
        signal_class_list.append(Sum_class(signal))
    return '初始化完成'


def start():
    '''
    主程序
    -----

    开始进行统计操作
    '''
    t0 = time.time()
    print('正在初始化...')
    init_signal_class()
    print('初始化完成!')
    print('Signal name', ' '*8, 'Profit(￥)\t\tVolume\tP/V')
    for signal in signal_class_list:
        signal.check()
        #print('[', signal.name, ']', ' '*(15-len(signal.name)),signal.profit)
        print('[', signal.name, ']', ' '*(15-len(signal.name)), round(signal.sum_profit, 2),
              '\t\t', signal.sum_volume, '\t', round(signal.sum_profit/signal.sum_volume, 3))
    t1 = time.time()
    tpy = t1-t0
    print('Finished Successfully in %5.3f seconds.' % tpy)


def profit_output():
    for i in range(len(signal_class_list)):
        p = signal_class_list[i].profit
        print(pd.DataFrame({'Profit': list(p.values())}, index=pd.MultiIndex.from_product(
            [[signal_class_list[i].name], p.keys()])))


def volume_output():
    for i in range(len(signal_class_list)):
        v = signal_class_list[i].volume
        print(pd.DataFrame({'Volume': list(v.values())}, index=pd.MultiIndex.from_product(
            [[signal_class_list[i].name], v.keys()])))


def output_to_excel():
    def beautify_excel():
        '''
        美化输出的excel文件
        '''
        from openpyxl.styles import Font, Border
        from openpyxl.formatting.rule import DataBarRule
        from openpyxl.drawing.image import Image

        # 打开文件，读取输出
        result = openpyxl.load_workbook(f[:-4]+' [OUTPUT].xlsx')
        table = result.worksheets[0]
        nrows = table.max_row

        # 准备样式（处理字体，取消边框）
        font = Font(name='dengxian', size=12)
        rule = DataBarRule(start_type='min', start_value=0, end_type='max', end_value=90,
                        color="FFFF0000", showValue="None", minLength=None, maxLength=None)

        #设置 Sheet name，添加列名
        table.title = '收益统计'
        table['A1'].value = 'Open Signal'
        table['B1'].value = 'Close Signal'    

        # 去除重复的列名，去除杂乱的边框
        for row in range(nrows+1, 0, -1):
            for j in range(len(table[row])):
                table[row][j].font = font
                table[row][j].border = Border(outline=False)
                if table[row][j].value == table[1][j].value and row > 1:
                    table[row][j].value = None

        # 加入数据条
        table.conditional_formatting.add('C1:C'+str(nrows), rule)

        # 设置列宽
        table.column_dimensions['A'].width = 13    
        table.column_dimensions['B'].width = 13
        table.column_dimensions['C'].width = 14

        # 插入图表
        wg = result.create_sheet(title='收益率平稳性')
        # wg.active
        graph = Image(f[:-4]+'AC figure.png')
        wg.append([' '])
        wg.add_image(graph,'A1')

        result.save(f[:-4]+' [OUTPUT].xlsx')
        print('已完成excel输出 文件路径 '+f[:-4]+' [OUTPUT].xlsx')
        return 0

    print("正在输出至excel文件...")
    writer = pd.ExcelWriter(f[:-4]+' [OUTPUT].xlsx')
    current_row = 0
    for i in range(len(signal_class_list)):
        #连结profit表和volume表
        v = signal_class_list[i].volume
        p = signal_class_list[i].profit
        w = signal_class_list[i].win
        l = signal_class_list[i].loss       
        wc = signal_class_list[i].win_count
        lc = signal_class_list[i].loss_count
        a = pd.DataFrame({'Profit': list(p.values())},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], p.keys()]))
        b = pd.DataFrame({'Volume': list(v.values())},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], v.keys()]))
        c = pd.DataFrame({'Win': list(w.values()),'Win Count': list(wc.values())},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], w.keys()]))
        d = pd.DataFrame({'Loss': list(l.values()),'Loss Count': list(lc.values())},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], l.keys()]))
        joint = a.join(b)
        joint['P/V'] = joint['Profit']/joint['Volume']

        joint = joint.join(c).join(d)
        joint = joint.fillna(0)
        joint['Win Ratio'] = joint['Win Count'] / (joint['Win Count']+joint['Loss Count'])
        joint['Expected Win'] = joint['Win'] #/ (joint['Win Count'] + joint['Loss Count'])
        joint['Expected Loss'] = joint['Loss']# / (joint['Win Count'] + joint['Loss Count'])
        joint = joint.fillna(0)
        joint['Expected Revenue'] = joint['Expected Win'] * joint['Win Ratio'] + joint['Expected Loss'] * (1-joint['Win Ratio'])

        joint = joint.drop(columns=['Win','Loss','Win Count','Loss Count'])
        joint.to_excel(writer, 'Sheet1', startrow=current_row)

        # 合并单元格
        # output_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row+len(v.keys()), end_column=1)
        current_row += len(v.keys())+1

    # 输出自相关系数曲线
    pd.plotting.autocorrelation_plot(list(t['% Profit'].dropna()), c='r').get_figure().savefig(f[:-4]+'AC figure.png')

    writer.save()
    # w.save(f+' output.xlsx')
    beautify_excel()    


if __name__ == '__main__':
    read_files(r"E:\RH\excel\I2015-2018_uncondition.xlsx")
    start()
    output_to_excel()
