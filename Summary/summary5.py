# -*- coding: utf-8 -*-
import time

import openpyxl
import pandas as pd
import numpy as np


def convert_types(table, nrows, ncols):
    global t

    def delete(original, pos):
        '''Inserts new inside original at pos.'''
        new = original[:pos] + original[pos+1:]
        return new

    def initial(sin):
        waiting = []
        count = 0
        for i in range(len(sin)):
            if sin[i] in ',¥':
                waiting.append(i)
        s = sin
        for i in waiting:
            s = delete(s, (i - count))
            count += 1
        return s

    abnormal = False
    for i in [4,7]:
        record = [0]
        for j in range(5, nrows+1):
            try:
                record.append(float(initial(table[j][i].value)))
            except:
                abnormal = True
                record.append(table[j][i].value)
        print('正在转换数据类型', t.keys()[i])
        t[t.keys()[i]] = record
    if abnormal:
        t.drop(t.index[[nrows-5, nrows-4]], inplace=True)  # 如果最后有未清仓 则删掉最后两行


class Sum_class():
    '''
    一个包含了signal的统计信息的类
    --------------------------

    包含的属性如下

    name, volume, profit

    需要逐个查看signal对象时，可以使用

    >>> s.signal_class_list
    '''

    def __init__(self, input_name):
        self.name = input_name
        self.volume = dict()
        self.sum_volume = 0
        self.profit = dict()
        self.sum_profit = 0
        self.timeseries = dict()
        self.win = dict(dict())  # 胜率
        # self.win_count = dict()
        self.loss = dict(dict())  # 赔率
        # self.loss_count = dict()
        self.max_drawdown = 0
        self.max_drawdown_date = 0

    def check(self):
        '''
        统计该Signal下的各类信息
        '''
        total_row = len(t)
        for i in range(total_row):
            is_active = t.loc[i]["类型"] in ["买入", "卖空"]  # 属于主动买入或者卖空操作 
                
            if t.loc[i]['Signal'] == self.name and is_active:
                # 记录volume
                j = i+1  # j为行控制变量，一直向下查找
                while t.loc[j]['#'] != t.loc[i]['#']+1:  # 当记录还没有跳到下一条时
                    # 记录总收益
                    self.sum_volume += t.loc[i]['Shares/Ctrts/Units - Profit/Loss']
                    self.volume[t.loc[j]['Signal']] = self.volume.get(
                        t.loc[j]['Signal'], 0) + t.loc[i]['Shares/Ctrts/Units - Profit/Loss']

                    # 记录总收益
                    self.sum_profit += t.loc[i]['Net Profit - Cum Net Profit']
                    self.profit[t.loc[j]['Signal']] = self.profit.get(
                        t.loc[j]['Signal'], 0) + t.loc[i]['Net Profit - Cum Net Profit']

                    # 记录胜率 赔率
                    if t.loc[i]['Net Profit - Cum Net Profit'] >= 0:
                        self.win[t.loc[j]['Signal']] = self.win.get(
                            t.loc[j]['Signal'], {})
                        self.win[t.loc[j]['Signal']][t.loc[j]['Date/Time']] = self.win[t.loc[j]['Signal']].get(
                            t.loc[j]['Date/Time'], 0)+t.loc[i]['Net Profit - Cum Net Profit']
                    else:
                        self.loss[t.loc[j]['Signal']] = self.loss.get(
                            t.loc[j]['Signal'], {})
                        self.loss[t.loc[j]['Signal']][t.loc[j]['Date/Time']] = self.loss[t.loc[j]['Signal']].get(
                            t.loc[j]['Date/Time'], 0)+t.loc[i]['Net Profit - Cum Net Profit']

                    # 记录收益对应的时间
                    self.timeseries[t.loc[j]['Signal']] = self.timeseries.get(
                        t.loc[j]['Signal'], {})
                    self.timeseries[t.loc[j]['Signal']][t.loc[j]['Date/Time']] = self.timeseries[t.loc[j]
                                                                                                 ['Signal']].get(t.loc[j]['Date/Time'], 0) + t.loc[i]['% Profit']
                    j += 1
                    if j >= total_row:
                        break

        # 胜率
        self.win_rate = len(self.win)/(len(self.win)+len(self.loss))


def read_files(filepath):
    '''
    t 为pandas载入的DataFrame 

    w 为openpyxl载入的原excel文件，用于转换数据类型

    f 为文件路径
    '''
    global t, w, f
    try:
        # 读取excel文件，sheet_name='交易列表',header=2（标题位于第三行）
        f = filepath
        t = pd.read_excel(filepath, sheet_name='交易列表', header=2, usecols='A:I' )
        
        w = openpyxl.load_workbook(filepath)
        table = w.worksheets[2]
        nrows = table.max_row
        ncols = table.max_column
    except:
        print('找不到文件')
    convert_types(table, nrows, ncols)
    t['Net Profit - Cum Net Profit']=t['Net Profit - Cum Net Profit']/200



def init_signal_class():
    '''
    获取所有主动操作的signal names。

    依次初始化各个signal:Sum_class对象。
    '''
    global signal_class_list

    # 统计所有属于主动买入的Signal类型
    buy_signal_types = t[t[u"类型"] == u"买入"].groupby('Signal')
    # 统计所有属于主动卖空的Signal类型
    short_signal_types = t[t['类型'] == '卖空'].groupby('Signal')
    buy_signals_name_list = buy_signal_types.size()._index
    short_signals_name_list = short_signal_types.size()._index

    signal_class_list = []
    for signal in buy_signals_name_list:
        signal_class_list.append(Sum_class(signal))
    for signal in short_signals_name_list:
        signal_class_list.append(Sum_class(signal))
    return '初始化完成'


def start():
    '''
    主程序
    -----

    开始进行统计操作
    '''
    global frequency

    def daily_frequency():
        frequency = dict()
        for i in range(1, len(t)):
            #统计每天次数
            date = t['Date/Time'][i].floor('D')
            frequency[date] = frequency.get(date, 0) + 1
        return frequency

    t0 = time.time()
    print('正在初始化...')
    init_signal_class()
    print('初始化完成!')
    print('Signal name', ' '*8, 'Profit(￥)\t\tVolume\tP/V')
    for signal in signal_class_list:
        signal.check()
        print('[', signal.name, ']', ' '*(15-len(signal.name)), round(signal.sum_profit, 2),
              '\t\t', signal.sum_volume, '\t', round(signal.sum_profit/signal.sum_volume, 3))

    frequency = pd.Series(daily_frequency())
    t1 = time.time()
    tpy = t1-t0
    print('Finished Successfully in %5.3f seconds.' % tpy)


def profit_output():
    for i in range(len(signal_class_list)):
        p = signal_class_list[i].profit
        print(pd.DataFrame({'Profit': list(p.values())}, index=pd.MultiIndex.from_product(
            [[signal_class_list[i].name], p.keys()])))


def volume_output():
    for i in range(len(signal_class_list)):
        v = signal_class_list[i].volume
        print(pd.DataFrame({'Volume': list(v.values())}, index=pd.MultiIndex.from_product(
            [[signal_class_list[i].name], v.keys()])))


def output_to_excel():
    def beautify_excel():
        '''
        美化输出的excel文件
        '''
        from openpyxl.styles import Font, Border, numbers
        from openpyxl.formatting.rule import DataBarRule
        from openpyxl.drawing.image import Image

        # 打开文件，读取输出
        result = openpyxl.load_workbook(f[:-4]+' [OUTPUT].xlsx')
        table = result.worksheets[0]
        nrows = table.max_row

        # 准备样式（处理字体，取消边框）
        font = Font(name='dengxian', size=12)
        rule = DataBarRule(start_type='min', start_value=0, end_type='max', end_value=90,
                           color="FFFF0000", showValue="None", minLength=None, maxLength=None)
        format_number = numbers.BUILTIN_FORMATS[40]
        format_percent = numbers.BUILTIN_FORMATS[10]

        # 设置 Sheet name，添加列名
        table.title = '收益统计'
        table['A1'].value = 'Open Signal'
        table['B1'].value = 'Close Signal'

        # 去除重复的列名，去除杂乱的边框
        for row in range(nrows+1, 0, -1):
            for j in range(len(table[row])):
                table[row][j].border = Border(outline=False)
                if table[row][j].value == table[1][j].value and row > 1:
                    table[row][j].value = None

        # 加入数据条
        table.conditional_formatting.add('C1:C'+str(nrows), rule)

        # 设置列宽
        table.column_dimensions['A'].width = 13
        table.column_dimensions['B'].width = 13
        table.column_dimensions['C'].width = 14
        table.column_dimensions['D'].width = 14
        table.column_dimensions['E'].width = 20
        table.column_dimensions['F'].width = 20        
        table.column_dimensions['G'].width = 10     
        table.column_dimensions['H'].width = 8      
        table.column_dimensions['I'].width = 9.5 
        table.column_dimensions['J'].width = 8                  
        table.column_dimensions['K'].width = 10
        table.column_dimensions['L'].width = 9                                          
        table.column_dimensions['M'].width = 13
        table.column_dimensions['N'].width = 13
        table.column_dimensions['O'].width = 16

        for c in ['E','H','G','J','M','N','O']:
            for irow in range(2,nrows+1):
                if table[c+str(irow)].value != None:
                    table[c+str(irow)].number_format = format_number
        for c in ['D', 'L']:
            for irow in range(2,nrows+1):
                if table[c+str(irow)].value != None:
                    table[c+str(irow)].number_format = format_percent
        #统计回撤
        def total_summary(t: pd.DataFrame):
            '''
            计算最大回撤比例,以及累计收益率

            返回结果：最大回撤率，开始日期，结束日期，总收益率
            '''
            t['Capital'] = float(t['价格'][1])+t['Net Profit - Cum Net Profit']

            
            yearly_drawdown = dict()
            
            t['Year']=pd.Series(t['Date/Time'][i].year for i in range(len(t)))
            t_group=t.groupby('Year')
            year_groups=[t_group.get_group(i) for i in t_group.groups.keys()]
            for year_group in year_groups:
                max_draw_down, temp_max_value = 0, 0
                start_date, end_date, current_start_date = 0, 0, 0
                continous = False # 是否连续
                for i in year_group.index:
                    if year_group['#'][i]>0: continue
                    if temp_max_value < year_group['Capital'][i]:
                        current_start_date = year_group['Date/Time'][i]
                    temp_max_value = max(temp_max_value, year_group['Capital'][i])
                    if max_draw_down>year_group['Capital'][i]/temp_max_value-1:
                        if not continous: 
                            continous = True
                        max_draw_down = year_group['Capital'][i]/temp_max_value-1
                    else:
                        if continous:
                            continous = False
                            start_date = current_start_date
                            end_date = year_group['Date/Time'][i]
                yearly_drawdown[year_group['Year'][i]] = max_draw_down, start_date, end_date            

            yearly_return = dict() # 记录年收益率
            max_draw_down, temp_max_value = 0, 0
            start_date, end_date, current_start_date = 0, 0, 0
            continous = False # 是否连续

            for i in range(2, len(t),2): # 偶数行的数据
                if temp_max_value < t['Capital'][i-2]:
                    current_start_date = t['Date/Time'][i]
                temp_max_value = max(temp_max_value, t['Capital'][i-2])
                if max_draw_down>t['Capital'][i]/temp_max_value-1:
                    if not continous: 
                        continous = True
                    max_draw_down = t['Capital'][i]/temp_max_value-1
                else:
                    if continous:
                        continous = False
                        start_date = current_start_date
                        end_date = t['Date/Time'][i]

                # 统计年收益率
                year = t['Date/Time'][i].year
                yearly_return[year] = t['Net Profit - Cum Net Profit'][i]
            
            total_return = t['Capital'][i]/t['Capital'][0]-1
            yearly_return = pd.Series(yearly_return) / t['Capital'][0]
            first_year = t['Date/Time'][1].year, yearly_return[t['Date/Time'][1].year]
            yearly_return = yearly_return.diff()
            yearly_return[first_year[0]] = first_year[1]
            return max_draw_down, start_date, end_date, total_return, pd.Series(yearly_return), yearly_drawdown

        temp = total_summary(t)
        table['C'+str(nrows+2)].value='最大回撤'
        table['D'+str(nrows+2)].value=temp[0]
        table['D'+str(nrows+2)].number_format = format_percent
        table['E'+str(nrows+2)].value=temp[1]
        table['F'+str(nrows+2)].value=temp[2]

        yearly_drawdown = temp[5];i=0
        for year in yearly_drawdown.keys():
            table[nrows+3+i][2].value = year
            table[nrows+3+i][3].value = yearly_drawdown[year][0]
            table[nrows+3+i][3].number_format = format_percent
            table[nrows+3+i][4].value = yearly_drawdown[year][1]        
            table[nrows+3+i][5].value = yearly_drawdown[year][2]       
            i+=1  
            
        
        table['A'+str(nrows+2)].value='累计收益率'
        table['B'+str(nrows+2)].value=temp[3]
        table['B'+str(nrows+2)].number_format = format_percent        

        yearly_return = temp[4];i=0
        for year in yearly_return.keys():
            table[nrows+3+i][0].value = year
            table[nrows+3+i][1].value = yearly_return[year]
            table[nrows+3+i][1].number_format = format_percent   
            i+=1        
        
        #统计交易次数
        def write_frequency(start_row, start_col):
            table[start_row-1][start_col].value = '每日交易次数'
            fd = frequency.describe()
            for i in range(len(fd)):
                table[start_row+i][start_col].value = fd.keys()[i]
                table[start_row+i][start_col+1].value = fd[i]
                table[start_row+i][start_col+1].number_format=format_number
        
        write_frequency(nrows+3, 6)

        for row in range(1,table.max_row+1):
            for j in range(0,table.max_column):
                try:
                    table[row][j].font = font
                except:pass
        # 插入图表
        wg = result.create_sheet(title='收益率平稳性')
        graph = Image(f[:-4]+'AC figure.png')
        wg.append([' '])
        wg.add_image(graph, 'A1')

        # 保存
        result.save(f[:-4]+' [OUTPUT].xlsx')         
        print('已完成excel输出 文件路径 '+f[:-4]+' [OUTPUT].xlsx')
        return 0

    def drawdown(t: pd.DataFrame):
        '''
        计算最大回撤比例
        '''
        t['Capital'] = (1+t[0]).cumprod()
        max_draw_down = 0
        temp_max_value = 0
        for i in range(1, len(t)):
            temp_max_value = max(temp_max_value, t['Capital'][i-1])
            max_draw_down = min(
                max_draw_down, t['Capital'][i]/temp_max_value-1)
        return max_draw_down

    def sharp(ts: pd.Series, risk_free=0.03):
        '''
        ts: Datetime Series

        用(年化收益率-无风险利率3%)/收益率标准差

        如果记录太少，则Sharp比率不具有参考意义
        '''
        ratio = (1+ts).cumprod()[-1]/(1+ts).cumprod()[0]  
        # 只考虑某种信号所造成的收益率：用最后的净值/最先的净值作为比例。(eg. 1.06/1.02)
        duration = (ts.index[-1]-ts.index[0]).days/365  
        # 从第一次交易到最后一次交易的时间
        if (duration == 0):
            yearly_rate = ts[0] 
            #如果只有一次交易，时间跨度是0不能计算，则用该交易的收益率作为yearly_rate
        else:
            yearly_rate = ratio**(1/duration)-1  # 算出年化收益率

        return (yearly_rate-risk_free)/ts.std()

    print("正在输出至excel文件...")
    writer = pd.ExcelWriter(f[:-4]+' [OUTPUT].xlsx')
    current_row = 0
    for i in range(len(signal_class_list)):
        # 连结profit表和volume表
        v = signal_class_list[i].volume
        p = signal_class_list[i].profit
        w = []
        wc = []
        w_cv = []
        l = []
        lc = []
        l_cv = []
        maxd = []
        s = []
        for item in (signal_class_list[i].win[ii] for ii in signal_class_list[i].win.keys()):
            w.append(sum(item.values()))
            wc.append(len(item))
            if wc[-1] != 0:
                w_cv.append(np.std(list(item.values())) / (w[-1] / wc[-1]))
            else:
                w_cv.append(0)
        for item in (signal_class_list[i].loss[ii] for ii in signal_class_list[i].loss.keys()):
            l.append(sum(item.values()))
            lc.append(len(item))
            if lc[-1] != 0:
                l_cv.append(np.std(list(item.values()))/(l[-1] / lc[-1]))
            else:
                l_cv.append(0)
        for item in (signal_class_list[i].timeseries[ii] for ii in signal_class_list[i].timeseries.keys()):
            maxd.append(drawdown(pd.DataFrame(item, index=range(1)).T))
            s.append(sharp(pd.Series(item)))
            
        
        a = pd.DataFrame({'Profit': list(p.values())},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], p.keys()]))
        b = pd.DataFrame({'Volume': list(v.values()), 'Max Drawdown': maxd, 'Sharp': s},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], v.keys()]))
        c = pd.DataFrame({'Win': w, 'Win Count': wc, 'Win CV': w_cv},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], signal_class_list[i].win.keys()]))
        d = pd.DataFrame({'Loss': l, 'Loss Count': lc, 'Loss CV': l_cv},
                         index=pd.MultiIndex.from_product([[signal_class_list[i].name], signal_class_list[i].loss.keys()]))
        joint = a.join(b)
        joint['P/V'] = joint['Profit']/joint['Volume']

        joint = joint.join(c).join(d)
        joint = joint.fillna(0)
        joint['Win Rate'] = joint['Win Count'] / (joint['Win Count']+joint['Loss Count'])
        joint['Expected Win'] = joint['Win'] / joint['Win Count']
        joint['Expected Loss'] = joint['Loss'] / joint['Loss Count']
        joint = joint.fillna(0)
        joint['Expected Revenue'] = joint['Expected Win'] * joint['Win Rate'] + joint['Expected Loss'] * (1-joint['Win Rate'])

        joint = joint.drop(columns=['Win', 'Loss'])
        joint.to_excel(writer, 'Sheet1', startrow=current_row)

        current_row += len(v.keys())+1

    # 输出自相关系数曲线
    pd.plotting.autocorrelation_plot(
        list(t['% Profit'].dropna()), c='r').get_figure().savefig(f[:-4]+'AC figure.png')

    writer.save()
    beautify_excel()


if __name__ == '__main__':
    read_files(r"E:\RH\excel\500F.xlsx")
    start()
    output_to_excel()
