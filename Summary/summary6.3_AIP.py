# -*- coding: utf-8 -*-
import time
import os
import openpyxl
import pandas as pd


def read_files(filepath, symbol = '', Type = 'index'):
    '''
    股票:Type='stock', 指数:Type='index', 期货:Type='future'
    
    t 为pandas载入的DataFrame 

    w 为openpyxl载入的原excel文件，用于转换数据类型

    f 为文件路径
    '''
    global t, w, f, folder, global_filepath,  global_symbol, global_type
    def convert_types(table, nrows, ncols):
        def delete(original, pos):
            '''Inserts new inside original at pos.'''
            new = original[:pos] + original[pos+1:]
            return new

        def initial(sin):
            waiting = []
            count = 0
            for i in range(len(sin)):
                if sin[i] in ',¥':
                    waiting.append(i)
            s = sin
            for i in waiting:
                s = delete(s, (i - count))
                count += 1
            return s

        for i in [4,7]:
            record = [0]
            for j in range(5, nrows+1):
                try:
                    record.append(float(initial(table[j][i].value)))
                except:
                    record.append(table[j][i].value)
            print('正在转换数据类型', t.keys()[i])
            t[t.columns[i]] = record
        
        # 转换 “开盘价” 为 当前日期
        i = 2
        record = [time.strftime('%Y/%m/%d')]
        print("正在填充缺失值为当前日期", time.strftime('%Y/%m/%d'))
        for j in range(5, nrows+1):
            try:
                if table[j][i].value=="开盘价":
                    record.append(time.strftime('%Y/%m/%d'))
                else:
                    record.append(table[j][i].value)
            except:
                record.append(table[j][i].value)
        t[t.columns[i]] = record
        t.astype(dtype={t.columns[i]:'datetime64'})
            
    try:
        # 读取excel文件，sheet_name='交易列表',header=2（标题位于第三行）
        f, global_filepath = filepath, filepath
        folder = f[:f.rfind(".xls")]+"\\"
        f = f[f.rfind("\\")+1:f.rfind(".xls")]
        if not os.path.exists(folder):
            os.makedirs(folder)
        t = pd.read_excel(global_filepath, sheet_name='交易列表', header=2, usecols='A:I' )
        w = openpyxl.load_workbook(global_filepath)
        table = w.worksheets[2]
        nrows = table.max_row
        ncols = table.max_column
    except:
        print('找不到文件/读取文件错误')
        return -1
    print('正在初始化...')
    convert_types(table, nrows, ncols)
    if symbol=='': 
        # 如果未指定代码，则从ts原始文件中选取
        print('Warning: symbol missing\nWill try symbol found in input file...')
        symbol = pd.read_excel(global_filepath, sheet_name=-1,header=2).columns[1]
        print('The default symbol found is \"{}\"'.format(symbol))
    if Type.endswith('future'):
        t['Net Profit - Cum Net Profit']/=200
    global_symbol, global_type = symbol, Type


def start():
    '''
    主程序
    -----
    '''
    global frequency
    def daily_frequency():
        global deal_list, date_list
        frequency, deal_list = dict(), list()
        date_list = t['Date/Time'].apply(lambda x:x.floor('D'))
        for i in range(1, len(t), 2):
            # 统计每天次数
            # Remark: ts中有些定投订单是拆分下单的，实际上依据策略，每周仅买入一单
            date = date_list[i]
            frequency[date] =  1
            #添加交易记录deal对象
            this_row = t.loc[i]; next_row = t.loc[i+1]
            deal_list.append(
                Deal(index=this_row['#'], dealtype = this_row['类型'],
                    start=this_row['Date/Time'], end= next_row['Date/Time'],
                    start_price=this_row['价格'], end_price=next_row['价格'],
                    start_signal=this_row['Signal'], end_signal=next_row['Signal'],
                    volume=this_row['Shares/Ctrts/Units - Profit/Loss']))
        print('添加交易记录对象...')
        return frequency

    t0 = time.time()
    frequency = pd.Series(daily_frequency())
    t1 = time.time()
    tpy = t1-t0
    print('Summation Finished Successfully in %5.3f seconds.' % tpy)
    

class Deal():
    def __init__(self, index, dealtype, start, end, start_signal, end_signal, volume, start_price, end_price):
        '''
        多头dealtype = 1， 空头dealtype = 0
        '''
        self.index = index
        if dealtype=='买入':self.type = 1
        else: self.type = -1
        self.start, self.end = start, end
        self.start_signal = start_signal
        self.end_signal = end_signal
        self.volume = volume
        self.start_price, self.end_price = start_price, end_price
        self.profit = (end_price-start_price) * volume * self.type
        self.floating = 0
        self.duration = self.end - self.start
        self.confirmed = False
        self.investment = self.start_price * self.volume
    def is_valid(self, today):
        return (today>=self.start) & (today<self.end)
    def confirm_profit(self):
        self.confirmed = True
    def cal_floating(self, close):
        self.floating = (close - self.start_price) * self.volume * self.type
        return self.floating

#%%
def 每日浮动收益统计(symbol:str, Type:str):
    from data_reader import get_index_day, get_stock_day, get_index_future_day, get_comm_future_day
    import pymssql

    start_date = str(date_list.iloc[1])[:10]
    end_date = str(date_list.iloc[-1])[:10]

    print("Fetching data from SQL database...")
    if Type=='stock':
        close = get_stock_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='index':
        close = get_index_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='index_future':
        close = get_index_future_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='comm_future':
        close = get_comm_future_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='coin':
        conn = pymssql.connect(server='192.168.0.28', port=1433, user='sa', password='abc123', database='rawdata',charset = 'GBK')
        SQL = "select * from coin.dbo.coin_1min where "+"date between "+"\'"+start_date+"\' and \'" + end_date+"\'"
        temp_data = pd.read_sql(SQL, conn)
        close = temp_data.set_index('date')
        close = close.resample('1D').last()
    elif Type[:3] == 'AIP':
        if Type[-5:] == 'stock':
            close = get_stock_day(symbol, start_date, end_date, freq = '1D')
        elif Type[-5:] == 'index':
            close = get_index_day(symbol, start_date, end_date, freq = '1D')
        else:raise TypeError('Type "{}" is NOT acceptable, please check your input.'.format(Type))
    else:
        raise TypeError('Type "{}" is NOT acceptable, please check your input.'.format(Type))

    close.index = close.index + pd.Timedelta(24, unit='h')
    record = []
    截止当日的累计盈利, comfirmed_profit = 0, 0
    for day in close.index:
        float_profit = 0
        for deal in deal_list:
            if deal.is_valid(day):
                float_profit += deal.cal_floating(close.loc[day]['sclose'])
                deal.floating_profit=deal.cal_floating(close.loc[day]['sclose'])
            elif day.date()==deal.end.date():
                #如果是当天结束的，当天确认收益
                deal.confirm_profit()
                comfirmed_profit+=deal.profit
                #deal_list.remove(deal)
        截止当日的累计盈利 = comfirmed_profit + float_profit
        #print(day, int(float_profit), int(comfirmed_profit), int(截止当日的累计盈利),sep='\t')
        record.append((day, float_profit, comfirmed_profit, 截止当日的累计盈利))
    ans=pd.DataFrame(record,columns=('date','floating_profit','comfirmed_profit','accumulated_profit'))
    ans=ans.set_index('date')
    if Type.endswith('future'):
        choice = input("You are using futures;\nDo you want to multiply amount by 200?\nInput 1 for YES, 0 for NO: ")
        if choice=='0': future_multiplier = 1
        else: future_multiplier = 200
        ans[['floating_profit','comfirmed_profit','accumulated_profit']]*=future_multiplier
    ans['当日盈亏']=ans.accumulated_profit.diff()
    return ans

#%%
def output_to_excel():
    from openpyxl.styles import Font, Border, numbers
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.drawing.image import Image

    # %% 统计回撤
    def drawdown_by_time(t: pd.DataFrame):
        '''
        计算内容：最大回撤比例，累计收益率  
        计算方式：单利  
        返回结果：最大回撤率，开始日期，结束日期，总收益率，年化收益，年化回撤  
        '''
        t['Capital'] = t['accumulated_profit']

        yearly_drawdown = dict()
        t['Date/Time'] = t.index
        t['Year']=t['Date/Time'].apply(lambda x:x.year)
        t_group=t.groupby('Year')
        year_groups=[t_group.get_group(i) for i in t_group.groups.keys()]
        for year_group in year_groups:
            max_draw_down, temp_max_value = 0, year_group['Capital'][0]
            start_date, end_date, current_start_date = 0, 0, 0
            continous = False # 是否连续
            for i in year_group.index:
                if temp_max_value < year_group['Capital'][i]:
                    current_start_date = year_group['Date/Time'][i]
                    temp_max_value = max(temp_max_value, year_group['Capital'][i])
                    continous = False
                else:
                    if max_draw_down>year_group['Capital'][i]/temp_max_value-1:
                        if not continous: 
                            continous = True
                        max_draw_down = year_group['Capital'][i]/temp_max_value-1
                    else:
                        if continous:
                            continous = False
                            start_date = current_start_date
                            end_date = year_group['Date/Time'][i]
            yearly_drawdown[year_group['Year'][i]] = max_draw_down, start_date, end_date            

        max_draw_down, temp_max_value = 0, 0
        start_date, end_date, current_start_date = 0, 0, 0
        continous = False # 是否连续

        for i in t.index:
            if temp_max_value < t['Capital'][i]:
                current_start_date = t['Date/Time'][i]
                temp_max_value = max(temp_max_value, t['Capital'][i])
                continous = False
            else:
                if max_draw_down>t['Capital'][i]/temp_max_value-1:
                    if not continous: 
                        continous = True
                    max_draw_down = t['Capital'][i]/temp_max_value-1
                else:
                    if continous:
                        continous = False
                        start_date = current_start_date
                        end_date = t['Date/Time'][i]

        return max_draw_down, start_date, end_date, yearly_drawdown
    
    # %% 统计定投收益率
    def yearly_return_AIP(t: pd.DataFrame):
        yearly_return = dict() # 记录年收益率
        for i in t.index:
            # 统计年收益率
            year = t['Date/Time'][i].year
            yearly_return[year] = t['accumulated_profit'][i]

        total_return = t['Capital'][i]/t['Capital'][0]-1
        yearly_return = pd.Series(yearly_return) / t['Capital'][0]
        first_year = t['Date/Time'][1].year, yearly_return[t['Date/Time'][1].year]
        yearly_return = yearly_return.diff()
        yearly_return[first_year[0]] = first_year[1]

        return total_return, pd.Series(yearly_return)
    #%%
    def sharp_ratio(ts: pd.Series, risk_free=0.03):
        '''
        ts: Datetime Series

        用(年化收益率-无风险利率3%)/收益率标准差

        如果记录太少，则Sharp比率不具有参考意义
        '''
        ratio = (1+ts).cumprod()[-1]/(1+ts).cumprod()[0]  
        # 只考虑某种信号所造成的收益率：用最后的净值/最先的净值作为比例。(eg. 1.06/1.02)
        duration = (ts.index[-1]-ts.index[0]).days/365  
        # 从第一次交易到最后一次交易的时间
        if (duration == 0):
            yearly_rate = ts[0] 
            #如果只有一次交易，时间跨度是0不能计算，则用该交易的收益率作为yearly_rate
        else:
            yearly_rate = ratio**(1/duration)-1  # 算出年化收益率

        return (yearly_rate-risk_free)/ts.std()

    def beautify_excel():
        '''
        美化输出的excel文件，并将最大回撤率等数据写入excel文件。
        '''
        global font, rule, format_number, format_percent

        # 准备样式（处理字体，取消边框）
        font = Font(name='dengxian', size=12)
        rule = DataBarRule(start_type='min', start_value=0, end_type='max', end_value=90,
                           color="FFFF0000", showValue="None", minLength=None, maxLength=None)
        format_number = numbers.BUILTIN_FORMATS[40]
        format_percent = numbers.BUILTIN_FORMATS[10]

    def write_yeild_and_drawdown(nrows):
        temp = drawdown_by_time(ans)
        table['C'+str(nrows+1)].value='最大回撤'
        table['D'+str(nrows+1)].value=temp[0]
        table['D'+str(nrows+1)].number_format = format_percent
        table['E'+str(nrows+1)].value=temp[1]
        table['F'+str(nrows+1)].value=temp[2]

        yearly_drawdown = temp[3];i=0
        for year in yearly_drawdown.keys():
            table[nrows+3+i][2].value = year
            table[nrows+3+i][3].value = yearly_drawdown[year][0]
            table[nrows+3+i][3].number_format = format_percent
            try:
                table[nrows+3+i][4].value = yearly_drawdown[year][1].strftime('%Y-%m-%d')        
                table[nrows+3+i][5].value = yearly_drawdown[year][2].strftime('%Y-%m-%d')
            except:
                pass
            i+=1  
            
        temp = yearly_return_AIP(ans)
        table['A'+str(nrows+1)].value='累计收益率'
        table['B'+str(nrows+1)].value=temp[0]
        table['B'+str(nrows+1)].number_format = format_percent        

        yearly_return = temp[1];i=0
        for year in yearly_return.keys():
            table[nrows+3+i][0].value = year
            table[nrows+3+i][1].value = yearly_return[year]
            table[nrows+3+i][1].number_format = format_percent   
            i+=1       
        rule1 = DataBarRule(start_type='percentile', start_value=0, end_type='percentile', end_value=99,
                           color="FFFF0000", showValue="None", minLength=None, maxLength=60)
        rule2 = DataBarRule(start_type='percentile', start_value=90, end_type='percentile', end_value=0,
                           color="FF22ae6b", showValue="None", minLength=None, maxLength=60)
        table.conditional_formatting.add('B{}:B{}'.format(nrows+3, nrows+1+i), rule1)
        table.conditional_formatting.add('D{}:D{}'.format(nrows+3, nrows+1+i), rule2)

        for c in ['A','B','C','D','E','F']:
            table.column_dimensions[c].width = 14
        
    def write_frequency(start_row, start_col):
        '''
        统计交易次数
        '''
        table.append(['交易频率统计'])
        fd = frequency.describe()
        fd = list(fd.iteritems())
        for i in range(len(fd)):
            table.append(list(fd[i]))
            table[start_row+i][start_col+1].number_format=format_number
        for row in range(1,table.max_row+1):
            for j in range(0,table.max_column):
                try:
                    table[row][j].font = font
                except:pass    

    def write_sheet2_daily_profit():
        '''插入日收益详情'''
        global ans

        ans = 每日浮动收益统计(symbol=global_symbol, Type=global_type)
        daysheet = result.create_sheet(title='日收益详情')
        daysheet.append(['date','浮动盈亏','已确认收益','累积收益（含浮动收益）','当日盈亏'])
        for row in ans.index:
            daysheet.append([row]+list(ans.loc[row]))
        daysheet.column_dimensions['A'].width = 21
        for row in range(1,daysheet.max_row+1):
            for j in range(0,daysheet.max_column):
                try:
                    daysheet[row][j].font = font
                except:pass
    def write_sheet3_daily_profit_graph():
        daygraphsheet = result.create_sheet(title='日收益图')
        def generate_profit_curve():
            import matplotlib.pyplot as plt
            fig = plt.figure()
            fig.set_size_inches(18, 12)

            ax = fig.add_subplot(211) 
            ax.plot(ans.index,ans['comfirmed_profit'], linewidth=2, label='确认收益')
            ax.plot(ans.index,ans['accumulated_profit'], linewidth=2, label='累积收益')
            ax.fill_between(ans.index,ans.accumulated_profit, y2=0, 
                            where=(ans.accumulated_profit<ans.accumulated_profit.shift(1))| \
                                    ((ans.accumulated_profit>ans.accumulated_profit.shift(-1))&(ans.accumulated_profit>=ans.accumulated_profit.shift(1))),
                            facecolor='grey',
                            alpha=0.3) 
            #最大回撤标注
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False
            ax.legend(fontsize=15)
            plt.grid()  

            bx = fig.add_subplot(212) 
            bx.bar(ans.index, ans.当日盈亏.where(ans.当日盈亏>0), 2, label='当日盈亏', color='red',alpha=0.8)
            bx.bar(ans.index, ans.当日盈亏.where(ans.当日盈亏<0), 2, label='当日盈亏', color='green',alpha=0.8)
            bx.legend(fontsize=15)
            plt.grid()  
            fig.savefig(folder+f+'Daily Profit.png', dpi=144, bbox_inches = 'tight')
        generate_profit_curve()
        graph = Image(folder+f+'Daily Profit.png')
        daygraphsheet.append([' '])
        daygraphsheet.add_image(graph, 'A1')
    def write_sheet4_autocorrelation_graph():
        # 输出自相关系数曲线并保存
        import matplotlib.pyplot as plt
        plt.figure()
        pd.plotting.autocorrelation_plot(
            ans['accumulated_profit'].pct_change().dropna(), c='r').get_figure().savefig(folder+f+'AC figure.png')
        # 插入图表
        wg = result.create_sheet(title='收益率平稳性')
        graph = Image(folder+f+'AC figure.png')
        wg.append([' '])
        wg.add_image(graph, 'A1')

    t0=time.time()
    print("正在输出至excel文件...")
    
    result = openpyxl.Workbook()
    table = result.active
    beautify_excel()
    write_sheet2_daily_profit()
    write_sheet3_daily_profit_graph()
    write_yeild_and_drawdown(1)
    write_frequency(1, 0)

    write_sheet4_autocorrelation_graph()
 
    # 保存
    result.save(folder+f+'[OUTPUT].xlsx')         
    print('已完成excel输出 文件路径 '+folder+f+'[OUTPUT].xlsx')
    t1 = time.time()
    tpy = t1-t0
    print('Excel Output Finished Successfully in %5.3f seconds.' % tpy)

#%%
if __name__ == '__main__':
    '''
    普通股:
    ·Type = 'stock'

    指数：
    ·Type = 'index'

    股指期货：
    · symbol必须从 {'IF':沪深300, 'IC':中证500, 'IH':上证50} 中选择
    · Type = 'index_future'

    商品期货：
    · symbol请查阅公司数据库手册-附录-商品代码
    · Type = 'comm_future'

    数字货币：
    · symbol = {'bitcoin':比特币}
    · Type = 'coin'

    定投:
    · symbol = {'AIP':定投(含有大量未平仓数据)} # 默认为定投某种指数，从excel文件中获取
    '''
    read_files(r"C:\Users\meiconte\Documents\RH\excel\TSoutput_stock\601888.SH.xlsx", symbol = '601888.SH', Type='AIP_stock') 
    start()
    output_to_excel()
