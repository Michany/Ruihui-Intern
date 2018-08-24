# -*- coding: utf-8 -*-
import time
import openpyxl
import pandas as pd


def read_files(filepath, symbol = '', Type = 'stock'):
    '''
    股票:Type='stock', 指数:Type='index', 期货:Type='future'
    
    t 为pandas载入的DataFrame 

    w 为openpyxl载入的原excel文件，用于转换数据类型

    f 为文件路径
    '''
    global t, w, f, folder, global_filepath,  global_symbol, global_type
    def convert_types(table, nrows, ncols):
        def delete(original, pos):
            '''Inserts new inside original at pos.'''
            new = original[:pos] + original[pos+1:]
            return new

        def initial(sin):
            waiting = []
            count = 0
            for i in range(len(sin)):
                if sin[i] in ',¥':
                    waiting.append(i)
            s = sin
            for i in waiting:
                s = delete(s, (i - count))
                count += 1
            return s

        for i in [4,7]:
            record = [0]
            for j in range(5, nrows+1):
                try:
                    record.append(float(initial(table[j][i].value)))
                except:
                    record.append(table[j][i].value)
            print('正在转换数据类型', t.keys()[i])
            t[t.columns[i]] = record
        
        # 转换 “开盘价” 为 当前日期
        i = 2
        record = [time.strftime('%Y/%m/%d')]
        print("正在填充缺失值为当前日期", time.strftime('%Y/%m/%d'))
        for j in range(5, nrows+1):
            try:
                if table[j][i].value=="开盘价":
                    record.append(time.strftime('%Y/%m/%d'))
                else:
                    record.append(table[j][i].value)
            except:
                record.append(table[j][i].value)
        t[t.columns[i]] = record
        t.astype(dtype={t.columns[i]:'datetime64'})
            
    try:
        # 读取excel文件，sheet_name='交易列表',header=2（标题位于第三行）
        f, global_filepath = filepath, filepath
        folder = f[:f.rfind(".xls")]+"\\"
        f = f[f.rfind("\\")+1:f.rfind(".xls")]
        t = pd.read_excel(global_filepath, sheet_name='交易列表', header=2, usecols='A:I' )
        w = openpyxl.load_workbook(global_filepath)
        table = w.worksheets[2]
        nrows = table.max_row
        ncols = table.max_column
    except:
        print('找不到文件/读取文件错误')
        return -1
    print('正在初始化...')
    convert_types(table, nrows, ncols)
    if symbol=='': 
        # 如果未指定代码，则从ts原始文件中选取
        print('Warning: symbol missing. Will try symbol found in input file...')
        symbol = pd.read_excel(global_filepath, sheet_name=-1,header=2).columns[1]
        print('The default symbol found is \"{}\"'.format(symbol))
    if Type.endswith('future'):
        t['Net Profit - Cum Net Profit']/=200
    global_symbol, global_type = symbol, Type


def start():
    '''
    主程序
    -----
    '''
    global frequency
    def daily_frequency():
        global deal_list, date_list
        frequency, deal_list = dict(), list()
        date_list = t['Date/Time'].apply(lambda x:x.floor('D'))
        for i in range(1, len(t), 2):
            # 统计每天次数
            # Remark: ts中有些定投订单是拆分下单的，实际上依据策略，每周仅买入一单
            date = date_list[i]
            frequency[date] =  1
            #添加交易记录deal对象
            this_row = t.loc[i]; next_row = t.loc[i+1]
            deal_list.append(
                Deal(index=this_row['#'], dealtype = this_row['类型'],
                    start=this_row['Date/Time'], end= next_row['Date/Time'],
                    start_price=this_row['价格'], end_price=next_row['价格'],
                    start_signal=this_row['Signal'], end_signal=next_row['Signal'],
                    volume=this_row['Shares/Ctrts/Units - Profit/Loss']))
        print('添加交易记录对象... (总计{}条交易)'.format(len(deal_list)))
        return frequency

    t0 = time.time()
    frequency = pd.Series(daily_frequency())
    t1 = time.time()
    tpy = t1-t0
    print('Summation Finished Successfully in %5.3f seconds.' % tpy)
    

class Deal():
    def __init__(self, index, dealtype, start, end, start_signal, end_signal, volume, start_price, end_price):
        '''
        多头dealtype = 1， 空头dealtype = 0
        '''
        self.index = index
        if dealtype=='买入':self.type = 1
        else: self.type = -1
        self.start, self.end = start, end
        self.start_signal = start_signal
        self.end_signal = end_signal
        self.volume = volume
        self.start_price, self.end_price = start_price, end_price
        self.profit = (end_price-start_price) * volume * self.type
        self.floating = 0
        self.duration = self.end - self.start
        self.confirmed = False 
        self.investment = self.start_price * self.volume
    def is_valid(self, today):
        return (today>=self.start) & (today<self.end)
    def confirm_profit(self):
        self.confirmed = True
    def cal_floating(self, close):
        self.floating = (close - self.start_price) * self.volume * self.type
        return self.floating

#%%
def 每日浮动收益统计(symbol:str, Type:str):
    from data_reader import get_index_day, get_stock_day, get_index_future_day, get_comm_future_day
    import pymssql

    start_date = str(date_list.iloc[1]-pd.Timedelta(30,'d'))[:10] # 早30天，为了使用RSI数据
    end_date = str(date_list.iloc[-1])[:10]
    print("[Time Range]",start_date,"~", end_date)
    
    print("Fetching data from SQL database...")
    if Type=='stock':
        close = get_stock_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='index':
        close = get_index_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='index_future':
        close = get_index_future_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='comm_future':
        close = get_comm_future_day(symbol, start_date, end_date, freq = '1D')
    elif Type=='coin':
        conn = pymssql.connect(server='192.168.0.28', port=1433, user='sa', password='abc123', database='rawdata',charset = 'GBK')
        SQL = "select * from coin.dbo.coin_1min where "+"date between "+"\'"+start_date+"\' and \'" + end_date+"\'"
        temp_data = pd.read_sql(SQL, conn)
        close = temp_data.set_index('date')
        close = close.resample('1D').last()
    elif Type == 'AIP':
        close = get_stock_day(symbol, start_date, end_date, freq = '1D')
    else:
        print('Type is NOT acceptable, please check your input.')
        quit()
    close.index = close.index + pd.Timedelta(24, unit='h')
    print("SQL data loaded...")
    
    print("统计当日盈亏中...")
    t0=time.time()
    record = []
    截止当日的累计盈利, comfirmed_profit = 0, 0
    for day in close.index[30:]: #close数据提前了30天
        float_profit, 累计投入 = 0, 0
        for deal in deal_list:
            if (deal.start<day) and (deal.end>day):
                累计投入 += deal.investment
            if deal.is_valid(day):
                float_profit += deal.cal_floating(close.loc[day]['sclose'])
                deal.floating_profit=deal.cal_floating(close.loc[day]['sclose'])
            elif day.date()==deal.end.date():
                #如果是当天结束的，当天确认收益
                deal.confirm_profit()
                comfirmed_profit+=deal.profit
                #deal_list.remove(deal)
        截止当日的累计盈利 = comfirmed_profit + float_profit
        #print(day, int(float_profit), int(comfirmed_profit), int(截止当日的累计盈利),sep='\t')
        record.append((day, float_profit, comfirmed_profit, 截止当日的累计盈利, 累计投入))
    ans=pd.DataFrame(record,columns=('date','floating_profit','comfirmed_profit','accumulated_profit','accumulated_investment'))
    ans=ans.set_index('date')
    if Type.endswith('future'):
        choice = input("You are using futures;\nDo you want to multiply amount by 200?\nInput 1 for YES, 0 for NO: ")
        if choice=='0': future_multiplier = 1
        else: future_multiplier = 200
        ans[['floating_profit','comfirmed_profit','accumulated_profit']]*=future_multiplier
    ans['当日盈亏']=ans.accumulated_profit.diff()
    t1=time.time()
    tpy = t1-t0
    print('统计当日盈亏已完成，用时 %5.3f 秒' % tpy)
    return ans, close

#%%
def AIP_net():
    '''
    查看定投某个指数得到的收益曲线
    '''
    global ans, stock_close
    temp = 每日浮动收益统计(symbol=global_symbol, Type=global_type)
    ans, stock_close = temp

def generate_profit_curve():
    import matplotlib.pyplot as plt
    print("定投收益曲线作图中...")
    t0=time.time()
    fig = plt.figure()
    fig.set_size_inches(12, 8)

    ax = fig.add_subplot(211) 
    ax.plot(ans.index,ans['comfirmed_profit'], linewidth=2, label='确认收益')
    ax.plot(ans.index,ans['accumulated_profit'], linewidth=2, label='累积收益')
    ax.fill_between(ans.index,ans.accumulated_profit, y2=0, 
                    where=(ans.accumulated_profit<ans.accumulated_profit.shift(1))| \
                            ((ans.accumulated_profit>ans.accumulated_profit.shift(-1))&(ans.accumulated_profit>=ans.accumulated_profit.shift(1))),
                    facecolor='grey',
                    alpha=0.3) 
    #最大回撤标注
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    ax.legend(fontsize=15)
    plt.grid()  

    bx = fig.add_subplot(212) 
    bx.bar(ans.index, ans.当日盈亏.where(ans.当日盈亏>0), 2, label='当日盈亏', color='red',alpha=0.8)
    bx.bar(ans.index, ans.当日盈亏.where(ans.当日盈亏<0), 2, label='当日盈亏', color='green',alpha=0.8)
    bx.legend(fontsize=15)
    plt.grid()  
    fig.savefig(folder[:-1]+' Daily Profit.png', dpi=144, bbox_inches = 'tight')
    t1=time.time()
    tpy = t1-t0
    print('收益曲线作图已完成，用时 %5.3f 秒' % tpy)

#%%
if __name__ == '__main__':
    '''
    普通股:
    ·Type = 'stock'

    指数：
    ·Type = 'index'

    股指期货：
    · symbol必须从 {'IF':沪深300, 'IC':中证500, 'IH':上证50} 中选择
    · Type = 'index_future'

    商品期货：
    · symbol请查阅公司数据库手册-附录-商品代码
    · Type = 'comm_future'

    数字货币：
    · symbol = {'bitcoin':比特币}
    · Type = 'coin'

    定投:
    · symbol = {'AIP':定投(含有大量未平仓数据)} # 默认为定投某种指数，从excel文件中获取
    '''
    read_files(r"C:\Users\meiconte\Documents\RH\excel\AIP.xlsx", symbol = '', Type='AIP') 
    start()
    AIP_net()
